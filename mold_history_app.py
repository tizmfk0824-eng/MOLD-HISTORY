import streamlit as st
import openpyxl
import warnings
warnings.filterwarnings('ignore')
from io import BytesIO
from copy import copy

st.set_page_config(page_title="금형 수리 이력 관리", page_icon="🔧", layout="centered")

st.title("🔧 금형 수리 이력 자동 입력")
st.markdown("---")

st.subheader("① 이력관리대장 파일 업로드")
mold_file = st.file_uploader("이력관리대장 (.xlsx)", type=["xlsx"], key="mold")

st.subheader("② 수리 입력 양식 업로드")
input_file = st.file_uploader("수리이력 입력양식 (.xlsx)", type=["xlsx"], key="input")

if mold_file and input_file:
    st.markdown("---")
    st.success("파일 업로드 완료! 아래 버튼을 클릭하세요.")

    if st.button("✅ 이력 자동 입력 실행", use_container_width=True):
        try:
            # 이력관리대장 열기
            wb_mold = openpyxl.load_workbook(BytesIO(mold_file.read()))
            ws_mold = wb_mold.active

            # 입력양식 열기
            wb_input = openpyxl.load_workbook(BytesIO(input_file.read()))
            ws_input = wb_input.active

            # 열 매핑 (입력양식 → 이력관리대장)
            # 입력양식: A=No, B=수리구분, C=발생일, D=금형/부품문제점, E=금형수리내역, F=완료일, G=수리결과, H=확인
            # 이력관리대장: A=No, C=수리구분, F=발생일, I=금형/부품문제점, Q=금형수리내역, Y=완료일, AB=수리결과, AE=확인
            col_map = {
                'B': 'C',   # 수리구분
                'C': 'F',   # 발생일
                'D': 'I',   # 금형/부품 문제점
                'E': 'Q',   # 금형 수리 내역
                'F': 'Y',   # 완료일
                'G': 'AB',  # 수리결과
                'H': 'AE',  # 확인
            }

            # 이력관리대장에서 다음 빈 행 찾기 (26행부터)
            start_row = 26
            next_row = start_row
            for r in range(start_row, ws_mold.max_row + 2):
                # No 열(A) 또는 발생일(F) 확인
                if ws_mold[f'F{r}'].value is None and ws_mold[f'I{r}'].value is None:
                    next_row = r
                    break

            # 입력양식에서 데이터 읽어서 입력
            inserted = 0
            for input_row in range(4, ws_input.max_row + 1):
                # 빈 행 스킵 (발생일이나 수리내역이 없으면)
                if ws_input[f'C{input_row}'].value is None and ws_input[f'D{input_row}'].value is None:
                    continue

                target_row = next_row + inserted

                # No 자동 입력
                # 이전 No 값 찾기
                prev_no = 0
                for r in range(target_row - 1, start_row - 1, -1):
                    if ws_mold[f'A{r}'].value and isinstance(ws_mold[f'A{r}'].value, (int, float)):
                        prev_no = int(ws_mold[f'A{r}'].value)
                        break
                ws_mold[f'A{target_row}'] = prev_no + 1 + inserted

                # 데이터 입력
                for input_col, mold_col in col_map.items():
                    value = ws_input[f'{input_col}{input_row}'].value
                    if value is not None:
                        ws_mold[f'{mold_col}{target_row}'] = value

                inserted += 1

            if inserted == 0:
                st.warning("입력양식에 데이터가 없어요. 내용을 작성 후 다시 업로드해주세요.")
            else:
                # 결과 파일 저장
                output = BytesIO()
                wb_mold.save(output)
                output.seek(0)

                st.success(f"✅ 총 {inserted}건 이력이 입력됐어요!")

                # 다운로드 버튼
                st.download_button(
                    label="📥 완성된 이력관리대장 다운로드",
                    data=output,
                    file_name=f"사출_금형_이력관리대장_업데이트.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )

        except Exception as e:
            st.error(f"오류가 발생했어요: {str(e)}")
            st.info("파일 형식이나 구조를 확인해주세요.")

else:
    st.info("📂 두 파일을 모두 업로드해주세요.")

st.markdown("---")
st.caption("금형 수리 이력 자동 입력 시스템 v1.0")
